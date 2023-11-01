import os
import time
import openpyxl
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import datetime
from datetime import timedelta
from openpyxl.utils import get_column_letter
import requests
import threading
import csv
import tkinter as tk
import threading
from tkinter import simpledialog
import subprocess
from tkinter import PhotoImage
from tkinter import ttk
import sys
import urllib.parse


def startDriver():
    """
    Starts Chrome webdriver
    """""
    newdriver = webdriver.Chrome()
    return newdriver


def startDashboard():
    """
    Sends an API request to 'https://verkehr.aachen.de'
    """""
    url = "https://verkehr.aachen.de/api/sensorthings/Things?$count=false&$filter=properties/type%20eq%20%27Verkehrszaehlstelle%27%20and%20properties/archive%20eq%20%27false%27&$expand=Locations,Datastreams(%24filter%3Dproperties%2FKlasse%20eq%20%27Bike%27%20and%20properties%2FAggregation%20eq%20%27d%27),Datastreams%2FObservedProperty,Datastreams%2FObservations(%24top%3D7%3B%24orderby%3DphenomenonTime%20desc%3B%24select%3Dresult%2CphenomenonTime%3B%24filter%3Ddate(phenomenonTime)%20ge%20date("+API_date+"))&$top=300&$select=@iot.id,description,name,properties/props&$orderBy=name"
    html_source = send_api_request_with_timeout(url)
    return html_source


def startWeather():
    """
    Returns the html source from 'https://www.wetter.com/wetter_aktuell/rueckblick/deutschland/aachen/DE0000003.html?timeframe=30d'
    """""
    url = 'https://www.wetter.com/wetter_aktuell/rueckblick/deutschland/aachen/DE0000003.html?timeframe=30d'
    response = requests.get(url)
    html_source = response.text
    return html_source


def startPictures():
    """
    Returns the html source from 'https://www.wetteronline.de/wetter/aachen'
    """""
    url = 'https://www.wetteronline.de/wetter/aachen'
    response = requests.get(url)
    html_source = response.text
    return html_source


def send_api_request_with_timeout(url, timeout=1, max_retries=10):
    for _ in range(max_retries):
        try:
            response = requests.get(url, timeout=timeout)
            response.raise_for_status()
            return response.text
        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}")
            continue
    raise Exception("Max retries reached, unable to fetch data from the API")


def getDataDashboard(content):
    """
    
    """  
    matches = re.findall('(.*?)\}\]\}\]\}', content)
    
    data_name_list = []
    data_list = []
    temp = []

    for match in matches:
        match1 = re.findall('label":"(.*?)"', match)
        match2 = re.findall('name":"(Rad[^"~]+)', match)
        match3 = re.findall('name":"Fahrr([^"~]+)', match)
        match4 = re.findall(str(date(2))+'T23:00:00\.000Z","result":\s*([0-9.]+)', match)
        
        if(len(match2) != 0):
            if(len(match2) > 2):
                temp = match2
                match2 = []
                match2.append(temp[0])
                match2.append(temp[1])

            temp = []
            temp.append(str(match1[0]) + " " + str(match2[0]))
            temp.append(str(match1[0])+ " " +str(match2[1]))
            data_name_list.append(temp)

        else:
            if(len(match3) > 2):
                    temp = match3
                    match3 = []
                    match3.append(temp[0])
                    match3.append(temp[1])

            temp = []
            temp.append(str(match1[0]) + " " + str(match3[0]))
            temp.append(str(match1[0])+ " " +str(match3[1]))
            data_name_list.append(temp)

        match4[0] = match4[0][:-2]
        if(len(match4) == 2):
            match4[1] = match4[1][:-2]
        data_list.append(match4)
             
    for data_name, data in zip(data_name_list, data_list):
        saveToExcel(data_name, data)
        csvParser(data_name)
        
    csvDataCollect(data_list, 2)


def getWeatherData(content):
    """
    Gathers all the needed weather data, parses the data to fit the needs
    """""
    regexes = [r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+,"precipitation":\d+\.\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+\.\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+,"precipitation":\d+\.\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+\.\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+,"precipitation":\d+\.\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+\.\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+,"precipitation":\d+\.\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+}',
               r'{"date":"'+date_reverse+'","windSpeed":\d+\.\d+,"temperatureMin":-?\d+\.\d+,"temperatureMax":-?\d+\.\d+,"precipitation":\d+\.\d+}']
    for regex in regexes:
        matches = re.findall(regex, content)
        if len(matches) != 0:
            break
    for match in matches:
        substring = match
        substring = substring.replace('"', '')
        substring = substring.replace('{', '')
        substring = substring.replace('}', '')
        datastring = substring.replace('date:', '')
        datastring = datastring.replace('windSpeed:', '')
        datastring = datastring.replace('temperatureMin:', '')
        datastring = datastring.replace('temperatureMax:', '')
        datastring = datastring.replace('precipitation:', '')
        allvalues = substring.split(',')
        datavalues = datastring.split(',')
        pattern = re.compile(r"(\w+):(\d+\.\d+)")
        dataNames = ["Min. Tempearatur in °C", "Max. Temperatur °C", "Niederschlagsmenge in l/m²"]
        del datavalues[0]
        del datavalues[0]
        for i, data in enumerate(datavalues):
            datavalues[i] = data.replace('.', ',')
        saveToExcel2(dataNames, datavalues)
        csvDataCollect(datavalues, 1)


def getBilder(content):
    """
    Gathers the name of the weather icon of date_today
    """""
    substringList = []
    trimmed_date = dateDEdate_Today[:-4]
    regex = r'<td class="" data-tt-args="\[&quot;'+trimmed_date+'&quot;,&quot;.*&quot;,&quot;.*&quot;,.,0, 0, &quot;&quot;, &quot;&quot;,0, &quot;&quot;, &quot;&quot;, &quot;&quot;, &quot;&quot;, &quot;&quot;, &quot;&quot;, &quot;&quot;, &quot;&quot;]" data-tt-function="TTwwsym">[\r\n]+ <img src="https:\/\/st\.wetteronline\.de\/dr\/1\.1\..*\/city\/prozess\/graphiken\/symbole\/standard\/farbe\/png\/[0-9][0-9]x[0-9][0-9]\/.*\.png'
    regexes = [r'bd____.png',
               r'bdg1__.png',
               r'bdg2__.png',
               r'bdgr1_.png',
               r'bdgr2_.png',
               r'bdr1__.png',
               r'bdr2__.png',
               r'bdr3__.png',
               r'bdsg__.png',
               r'bdsn1_.png',
               r'bdsn2_.png',
               r'bdsn3_.png',
               r'bdsr1_.png',
               r'bdsr2_.png',
               r'bdsr3_.png',
               r'nb____.png',
               r'ns____.png',
               r'so____.png',
               r'wb____.png',
               r'wbg1__.png',
               r'wbg2__.png',
               r'wbs1__.png',
               r'wbs2__.png',
               r'wbsg__.png',
               r'wbsns1.png',
               r'wbsns2.png',
               r'wbsrs1.png',
               r'wbsrs2.png',
               r'wbr1__.png',
               r'wbr2__.png',
               r'bw____.png',
               r'bws2__.png',
               r'bwr1__.png',
               r'bwr2__.png',
               r'bwg1__.png',
               r'bws1__.png']
    matches = re.findall(regex, content)
    match = ', '.join(matches)
    for regex in regexes:
        substrings = re.findall(regex, match)
        if len(substrings) != 0:
            break
    substring = ', '.join(substrings)
    if (not substring):
        saveToExcel3(match)
        substringList.append(match)
    else:
        saveToExcel3(substring)
        substringList.append(substring)
    csvDataCollect(substringList, 0)


def startExcel():
    """
    Loads/Creates an excel file named data.xlsx
    """""
    if not os.path.exists(current_path+"/Excel/data.xlsx"):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(current_path+"/Excel/data.xlsx")
    return workbook


def counter2():
    """
    Increments the counter by two and returns it
    """""
    counter2.count += 2
    return counter2.count


def counter2minus1():
    """
    Decrements the counter by 1
    """""
    counter2.count -= 1


def date(minus_days):
    """
    Gathers and returns the date from yesterday in format MM/DD/YYYY
    """""
    date_today = datetime.date.today()
    yesterday = date_today - timedelta(days=minus_days)
    date_str = str(yesterday)
    return date_str


def get_dateDE():
    """
    Gathers and returns the date from yesterday in format DD/MM/YYYY
    """""
    date_today = datetime.date.today()
    yesterday = date_today - timedelta(days=1)
    date_str = yesterday.strftime("%d.%m.%Y")
    return date_str


def dateDEdate_Today():
    """
    Gathers and returns the date from date_today in format DD/MM/YYYY
    """""
    date_today = datetime.date.today()
    date_str = date_today.strftime("%d.%m.%Y")
    return date_str


def get_API_date():
    from datetime import datetime, timedelta
    current_datetime = datetime.utcnow() - timedelta(days=7) 
    iso8601_format = current_datetime.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    url_encoded_iso8601_format = urllib.parse.quote(iso8601_format)
    return url_encoded_iso8601_format


def findFirstEmptyCol():
    """
    Returns the first empty column that is not the first one
    """""
    if not os.path.exists(current_path+"/Excel/data.xlsx"):
        return 1
    else:
        try:
            path = current_path+"/Excel/data.xlsx"
            df = pd.read_excel(path, sheet_name=0, engine='openpyxl')
            empty_col = df.iloc[:, 1:].columns[(df.iloc[:, 1:].isna().all())]
            if empty_col.empty:
                number = df.shape[1]
                return number
            else:
                return empty_col[0]
        except:
            print("find first empty col")


def findLastSavedDate():
    """
    Finds the date from the last time the script ran, returns true if the last saved date is yesterdays date. Returns false if that is not the case
    """""
    from datetime import datetime
    if not os.path.exists(current_path+"/Excel/data.xlsx"):
        return True
    else:
        try:
            path = current_path+"/Excel/data.xlsx"
            df = pd.read_excel(path)
            highest_column_value = df.iloc[0, df.shape[1] - 1]
            
            date_format = "%d.%m.%Y"
            formatted_highest_column_value = datetime.strptime(highest_column_value, date_format)
            formatted_dateDE = datetime.strptime(dateDE, date_format)
            
            if formatted_highest_column_value < formatted_dateDE:
                return True
            else:
                return False
        except Exception as e: 
            print(e)
            return True
           

def saveToExcel(filenames, data):
    """
    Writes the data gathered about the zählstellen and yesterdays date into the excel file
    """""
    worksheet.cell(row=2, column=col_num+2).value = dateDE
    row_num = counter2()
    
    worksheet.cell(row=row_num+4, column=col_num+1).value = (filenames[0]+":")
    worksheet.cell(row=row_num+4, column=col_num+2).value = int(data[0])
    if len(data) == 1:
        counter2minus1()
    if len(data) == 2:
        worksheet.cell(row=row_num+5, column=col_num+1).value = (filenames[1]+":")
        worksheet.cell(row=row_num+5, column=col_num+2).value = int(data[1])
        

def saveToExcel2(dataNames, datavalues):
    """
    Writes the weather data into the excel file
    """""
    worksheet.cell(row=4, column=col_num+1).value = dataNames[0]
    worksheet.cell(row=5, column=col_num+1).value = dataNames[1]
    worksheet.cell(row=6, column=col_num+1).value = dataNames[2]
    worksheet.cell(row=4, column=col_num+2).value = datavalues[0]
    worksheet.cell(row=5, column=col_num+2).value = datavalues[1]
    if datavalues[2] != "0":
        worksheet.cell(row=6, column=col_num+2).value = datavalues[2]


def saveToExcel3(imgName):
    """
    Writes the name of the weather icon to the excel file
    """""
    from datetime import datetime
    global dateDE
    temp1 = datetime.strptime(dateDE, '%d.%m.%Y')
    temp2 = datetime.strptime(dateDEdate_Today, '%d.%m.%Y')
    worksheet.cell(row=3, column=col_num+1).value = 'Wetter Symbol'
    if (not imgName or temp1 + timedelta(days=1) != temp2):
        worksheet.cell(row=3, column=col_num+2).value = "No Data"
    else:
        worksheet.cell(row=3, column=col_num+2).value = imgName


def checker():
    """
    Checks if all for the day data has been written into the excel file
    """""
    column_letter = openpyxl.utils.get_column_letter(col_num+1)
    column_index = openpyxl.utils.column_index_from_string(column_letter)
    column_1_empty = True
    for row in range(3, worksheet.max_row + 1):
        if (row != 6):
            cell_value1 = worksheet.cell(row=row, column=column_index).value
            if cell_value1 is None:
                column_1_empty = False
    return column_1_empty


def csvDataCollect(data_list, index):
    """
    Collects all data into a list
    """""
    listDateDE = []
    listDateDE.append(dateDE)
    csvDataLists[0] = listDateDE
    if index == 0:
        csvDataLists[1] = data_list
    elif index == 1:
        csvDataLists[2] = data_list
    elif index == 2:
        csvDataLists[3] = data_list
    

def csvParser(lines):
    """
    Parses zählstellen data into one list
    """""
    if (isinstance(lines, int)):
        dataList.append(lines)
    elif (len(lines) == 4):
        dataList.append(lines[1])
        dataList.append(lines[3])
    elif (len(lines) == 2):
        dataList.append(lines[1])
    

def csvBackup():
    """
    Writes all data into a new csv file for each day
    """""
    x = 0
    text = []
    text.append("Bilder Name nicht passend zum Datum +1 Tag")
    csvDataLists.insert(0, text)
    csvFileName = current_path+"/Csv/"+dateDE+".csv"
    if os.path.isfile(csvFileName):
        os.remove(csvFileName)
    with open(csvFileName, 'a', newline='') as file:
        writer = csv.writer(file)
        for csvDataList in csvDataLists:
            for data in csvDataList:
                x = x+1
                if (x == 3 and not data):
                    writer.writerow(["No Data"])
                else:
                    if (data != 0):
                        writer.writerow([data])
                    else:
                        writer.writerow(["No Data"])


current_path = "C:\\Users\\Fabian\\Desktop\\radwegzaehler"
# assign the value 1 to the count attribute of the counter2 object
counter2.count = 1
# create a date object representing the current date of yesterday
date_reverse = date(1)
# create a date object representing the current date in the German language of yesterday
dateDE = get_dateDE()
# create a date object representing the current date in the German language
dateDEdate_Today = dateDEdate_Today()
# find the first empty column in the worksheet and assign the result to the col_num variable
col_num = findFirstEmptyCol()
# create a new workbook object
workbook = startExcel()
# get the active worksheet in the workbook
worksheet = workbook.active
# 2d array for gathering data required for the csv backup
csvDataLists = [[], [], [], []]
# array for gathering data required for the csv backup
dataList = []
# position in px that the mouse gets moved during scraping process
move_mouse_x_koord = 1880
# temp storage
old_dates = [] 
old_dataList = []
xpath_cache = {}
API_date = get_API_date()

def scrape():
    max_retries = 5
    global col_num
    col_num = findFirstEmptyCol()
    try:
        # if the script has not run already date_today
        if findLastSavedDate():
            for _ in range(max_retries):
                counter2.count = 1
                # multithreading for the gathering and saving of the data from the three websites
                thread1 = threading.Thread(target=lambda: getDataDashboard(startDashboard()))
                thread2 = threading.Thread(target=lambda: getWeatherData(startWeather()))
                thread3 = threading.Thread(target=lambda: getBilder(startPictures()))
                thread4 = threading.Thread(target=lambda: csvBackup())
                thread5 = threading.Thread(target=lambda: workbook.save(current_path+"\Excel\data.xlsx"))
                thread1.start()
                thread2.start()
                thread3.start()
                threads = [thread1, thread2, thread3]
                for thread in threads:
                    thread.join()
                thread4.start()
                thread5.start()
                thread4.join()
                thread5.join()
                # if the excel file has been appropriately filled with data the loop breaks
                if (checker()):
                    break
    except Exception as e: 
        print("scrape failed")
        print(e)

    
def delete():
    """
    Will delete the last data entry
    """""
    if(findFirstEmptyCol()-1 >= 0):
        worksheet.delete_cols(findFirstEmptyCol()-1, 2)
        workbook.save(current_path+"/Excel/data.xlsx")


def pre_scrape(user_input):
    from datetime import datetime
    global dateDE
    dateDE = get_dateDE()
    date_format = '%d.%m.%Y'
    
    global date_reverse
    date_reverse = date(int(user_input))
    original_sequence = [6, 5, 4, 3, 2, 1]
    
    temp_date = datetime.strptime(dateDE, date_format)
    new_date = temp_date - timedelta(days=int(user_input)-1)
    dateDE = new_date.strftime(date_format)
    scrape()
    
    
def past_day():
    """
    Will run the scraper with the number of days in the past specified
    """""
    user_input = simpledialog.askstring("Input", "Enter a number:", parent=root)
    pre_scrape(user_input)
    
    
def file_location():
    root.attributes("-topmost", 0)
    subprocess.Popen(['explorer', current_path+"\\Excel\\"])
    
    
def run_action():
    """
    Runs the standard data scraper for yesterday
    """""
    def run():
        root.attributes("-topmost", 1)
        root.config(cursor="wait")
        button_run.config(bg="yellow")
        pre_scrape(1)
        button_run.config(bg="SystemButtonFace")
        root.config(cursor="arrow")
    thread = threading.Thread(target=run)
    thread.start()


def delete_action():
    """
    Will delete the last data entry and will make the button appear yellow and 
    set the cursor to the waiting icon for the duration of the deletion proses 
    """""
    def run():
        root.attributes("-topmost", 1)
        root.config(cursor="wait")
        button_delete.config(bg="yellow")
        delete()
        button_delete.config(bg="SystemButtonFace")
        root.config(cursor="arrow")
    thread = threading.Thread(target=run)
    thread.start()


def run_yesterday_action():
    """
    Will run the scraper with the number of days in the past specified
    and will make the button appear yellow and set the cursor to the waiting 
    icon for the duration of the deletion proses 
    """""
    root.attributes("-topmost", 1)
    root.config(cursor="wait")
    button_yesterday.config(bg="yellow")
    past_day()
    button_yesterday.config(bg="SystemButtonFace")
    root.config(cursor="arrow")
    
    
def open_file_location_action():
    root.config(cursor="wait")
    button_file_explorer.config(bg="yellow")
    file_location()
    button_file_explorer.config(bg="SystemButtonFace")
    root.config(cursor="arrow")


def setup_table():
    table_frame = tk.Frame(root)
    table_frame.place(y=60, relx=0.5, rely=0.6, anchor=tk.CENTER, width=650, height=130)
    
    global table
    table = ttk.Treeview(table_frame, columns=("1","2","3","4","5","6",), show="headings")

    col_width = 108
    table.column("1", anchor=tk.W, width=col_width)
    table.column("2", anchor=tk.W, width=col_width)
    table.column("3", anchor=tk.W, width=col_width)
    table.column("4", anchor=tk.W, width=col_width)
    table.column("5", anchor=tk.W, width=col_width)
    table.column("6", anchor=tk.W, width=col_width)
    
    def run():
        while(True):
            update_table()
            time.sleep(1)
    thread = threading.Thread(target=run)
    thread.start()
    

def read_excel():
    """
    Reads in data from the excel file to display it in the GUI
    """""
    dates = []
    images = []
    weatherList = []
    dataList = []
    col_num = findFirstEmptyCol()
    if os.path.exists(current_path+"/Excel/data.xlsx"):
        for x in range(0,6):
            if(col_num-x*2 > 0):
                column_letter = openpyxl.utils.get_column_letter(col_num-x*2)
            else:
                column_letter = "A"
            column_index = openpyxl.utils.column_index_from_string(column_letter)
            
            date = worksheet.cell(row=2, column=column_index).value
            dates.append(date)
            
            image = worksheet.cell(row=3, column=column_index).value
            images.append(image)
            
            for row in range(4, 7):
                weather = worksheet.cell(row=row, column=column_index).value
                weatherList.append(weather)

            x = 0
            tempList = []
            
            for row in range(7, worksheet.max_row + 1):
                    x += 1
                    cell_value = worksheet.cell(row=row, column=column_index).value
                    if cell_value is not None:
                        tempList.append(True)
                        
            if len(tempList) != x:
                dataList.append(False)
            else:
                dataList.append(True)
    else:
        dates = [0,0,0,0,0,0]
        images = [0,0,0,0,0,0]
        weatherList = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        dataList = [False,False,False,False,False,False]
    return dates, images, weatherList, dataList
        

def update_table():
    global old_dates
    global old_dataList
    
    dates, images, weather, dataList = read_excel()

    if(old_dates != dates or dataList != old_dataList):
        table.delete(*table.get_children())

        table.heading("1", text=dates[5])
        table.heading("2", text=dates[4])
        table.heading("3", text=dates[3])
        table.heading("4", text=dates[2])
        table.heading("5", text=dates[1])
        table.heading("6", text=dates[0])

        data = [
            (images[5],images[4],images[3],images[2],images[1],images[0]),
            (weather[15],weather[12],weather[9],weather[6],weather[3],weather[0]),
            (weather[16],weather[13],weather[10],weather[7],weather[4],weather[1]),
            (weather[17],weather[14],weather[11],weather[8],weather[5],weather[2]),
            (dataList[5],dataList[4],dataList[3],dataList[2],dataList[1],dataList[0],)
        ]

        for item in data:
            table.insert("", tk.END, values=item)

        table.pack()
        old_dates = dates
        old_dataList = dataList


def setup():
    """
    Setup for the GUI
    """""
    root.title("Data Scraper")

    icon_path = "krackenIcon.ico"
    root.iconbitmap(get_path(icon_path))

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_position = (screen_width - window_width) // 2
    y_position = (screen_height - window_height) // 2 - -50

    root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
    
    background_image = PhotoImage(file=get_path("krackenIcon.png"))
    background_label = tk.Label(root, image=background_image)
    background_label.place(relwidth=1, relheight=1)  

    labels()
    buttons()
    setup_table()
    
    root.mainloop()
    

def labels():
    """
    Creates all three labels and gives them their text, sizes and positions
    """""
    label_width = 25
    label_height = 2
    label1 = tk.Label(root, text="Scrapes data from yesterday", width=label_width, height=label_height, wraplength=200)
    label1.place(x=window_width/2-window_width/3-label1.winfo_reqwidth()/2, y=175)

    label2 = tk.Label(root, text="Deletes the last data entry", width=label_width, height=label_height, wraplength=200)
    label2.place(x=window_width/2-label2.winfo_reqwidth()/2, y=175)

    label3 = tk.Label(root, text="Choose from how many days ago the data gets scraped (1-6)", width=label_width, height=label_height, wraplength=200)
    label3.place(x=window_width/2+window_width/3-label3.winfo_reqwidth()/2, y=175)
    
    label3 = tk.Label(root, text="Opens the file explorer", width=label_width, height=label_height, wraplength=200)
    label3.place(x=window_width/2-label2.winfo_reqwidth()/2, y=60)


def buttons():
    """
    Creates all three buttons and gives them their text, sizes and positions
    """""
    button_width = 25
    global button_run, button_delete, button_yesterday, button_file_explorer
    
    button_run = tk.Button(root, text="RUN", width=button_width, height=2, command=run_action)
    button_run.place(x=window_width/2-window_width/3-button_run.winfo_reqwidth()/2, y=125)

    button_delete = tk.Button(root, text="DELETE", width=button_width, height=2, command=delete_action)
    button_delete.place(x=window_width/2-button_delete.winfo_reqwidth()/2, y=125)

    button_yesterday = tk.Button(root, text="RUN PAST DAYS", width=button_width, height=2, command=run_yesterday_action)
    button_yesterday.place(x=window_width/2+window_width/3-button_yesterday.winfo_reqwidth()/2, y=125)
    
    button_file_explorer = tk.Button(root, text="OPEN FILE LOCATION", width=button_width, height=2, command=open_file_location_action)
    button_file_explorer.place(x=window_width/2-button_delete.winfo_reqwidth()/2, y=10)


def get_path(filename):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, filename)
    else:
        return filename


root = tk.Tk()
# Width and height of the GUI window
window_width = 700
window_height = 400
# Makes sure that the GUI stays on top and doesn't get minimized
root.attributes("-topmost", 1)
setup()